"""Tests for Worklog Excel export API — See VRIP-132 AC."""

import uuid
from datetime import date, timedelta
from io import BytesIO

import pytest
from httpx import AsyncClient
from openpyxl import load_workbook
from sqlalchemy.ext.asyncio import AsyncSession

from app.modules.allocations.models import Assignment
from app.modules.clients.models import Client
from app.modules.projects.models import Project
from app.modules.resources.models import Resource
from app.modules.worklogs.models import Worklog
from tests.conftest import create_test_user, login_as_role

XLSX_CT = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


async def _seed_resource(db: AsyncSession, name: str) -> Resource:
    r = Resource(
        id=uuid.uuid4(),
        employee_id=f"EMP-{uuid.uuid4().hex[:6]}",
        name=name,
        designation="Developer",
        date_of_joining=date.today() - timedelta(days=90),
        is_active=True,
    )
    db.add(r)
    await db.flush()
    return r


async def _seed_client(db: AsyncSession) -> Client:
    c = Client(id=uuid.uuid4(), name=f"Client-{uuid.uuid4().hex[:4]}", is_active=True)
    db.add(c)
    await db.flush()
    return c


async def _seed_project(db: AsyncSession, client_id, dm_id, pm_id, **kwargs) -> Project:
    p = Project(
        id=uuid.uuid4(),
        name=kwargs.get("name", f"Proj-{uuid.uuid4().hex[:4]}"),
        client_id=client_id,
        dm_id=dm_id,
        pm_id=pm_id,
        worklog_enabled=True,
        status="ACTIVE",
        is_active=True,
    )
    db.add(p)
    await db.flush()
    return p


async def _seed_assignment(db: AsyncSession, project_id, resource_id) -> Assignment:
    a = Assignment(
        id=uuid.uuid4(),
        project_id=project_id,
        resource_id=resource_id,
        allocation_pct=100,
        billability_pct=100,
        start_date=date.today() - timedelta(days=60),
        status="ACTIVE",
    )
    db.add(a)
    await db.flush()
    return a


async def _seed_worklog(db: AsyncSession, resource_id, project_id, log_date, hours=4.0) -> Worklog:
    w = Worklog(
        id=uuid.uuid4(),
        resource_id=resource_id,
        project_id=project_id,
        log_date=log_date,
        hours=hours,
        note=f"Work on {log_date}",
    )
    db.add(w)
    await db.flush()
    return w


async def _setup_data(db: AsyncSession):
    """Seed two engineers, one project, and worklogs."""
    dm = await _seed_resource(db, "DM User")
    eng1 = await _seed_resource(db, "Alice")
    eng2 = await _seed_resource(db, "Bob")
    cl = await _seed_client(db)
    project = await _seed_project(db, cl.id, dm.id, dm.id, name="TestProject")
    await _seed_assignment(db, project.id, eng1.id)
    await _seed_assignment(db, project.id, eng2.id)
    yesterday = date.today() - timedelta(days=1)
    two_days_ago = date.today() - timedelta(days=2)
    await _seed_worklog(db, eng1.id, project.id, yesterday, 4.0)
    await _seed_worklog(db, eng1.id, project.id, two_days_ago, 3.0)
    await _seed_worklog(db, eng2.id, project.id, yesterday, 5.0)
    await db.commit()
    return dm, eng1, eng2, project


def _parse_xlsx(content: bytes):
    wb = load_workbook(BytesIO(content))
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    headers = rows[0] if rows else []
    data = rows[1:] if len(rows) > 1 else []
    return headers, data


# ──────────────────────────────────────────────
# GET /api/v1/worklogs/export — Company-wide
# ──────────────────────────────────────────────


@pytest.mark.asyncio
async def test_export_worklogs_ceo(client: AsyncClient, db: AsyncSession):
    dm, eng1, eng2, project = await _setup_data(db)
    await login_as_role(client, db, "CEO")

    resp = await client.get("/api/v1/worklogs/export")
    assert resp.status_code == 200
    assert resp.headers["content-type"] == XLSX_CT
    assert "attachment" in resp.headers["content-disposition"]
    assert ".xlsx" in resp.headers["content-disposition"]

    headers, data = _parse_xlsx(resp.content)
    assert headers == ("Date", "Resource", "Project", "Hours", "Note")
    assert len(data) == 3


@pytest.mark.asyncio
async def test_export_worklogs_with_date_filter(client: AsyncClient, db: AsyncSession):
    dm, eng1, eng2, project = await _setup_data(db)
    await login_as_role(client, db, "CEO")

    yesterday = (date.today() - timedelta(days=1)).isoformat()
    resp = await client.get(f"/api/v1/worklogs/export?start_date={yesterday}&end_date={yesterday}")
    assert resp.status_code == 200
    _, data = _parse_xlsx(resp.content)
    assert len(data) == 2  # only yesterday's entries (eng1 + eng2)
    assert yesterday in resp.headers["content-disposition"]


@pytest.mark.asyncio
async def test_export_worklogs_empty_result(client: AsyncClient, db: AsyncSession):
    await login_as_role(client, db, "CEO")
    resp = await client.get("/api/v1/worklogs/export")
    assert resp.status_code == 200
    headers, data = _parse_xlsx(resp.content)
    assert headers == ("Date", "Resource", "Project", "Hours", "Note")
    assert len(data) == 0


@pytest.mark.asyncio
async def test_export_worklogs_engineer_self_only(client: AsyncClient, db: AsyncSession):
    """ENGINEER with SELF_ONLY scope sees only own worklogs in export."""
    dm, eng1, eng2, project = await _setup_data(db)
    user = await create_test_user(db, "ENGINEER", name="Eng1 Export")
    user.resource_id = eng1.id
    await db.commit()
    await client.post("/api/v1/auth/login", json={"email": user.email, "password": "TestPass123"})

    resp = await client.get("/api/v1/worklogs/export")
    assert resp.status_code == 200
    _, data = _parse_xlsx(resp.content)
    assert len(data) == 2  # only eng1's 2 entries


@pytest.mark.asyncio
async def test_export_worklogs_unauthenticated(client: AsyncClient):
    resp = await client.get("/api/v1/worklogs/export")
    assert resp.status_code == 401


# ──────────────────────────────────────────────
# GET /api/v1/worklogs/my/export — Personal
# ──────────────────────────────────────────────


@pytest.mark.asyncio
async def test_export_my_worklogs(client: AsyncClient, db: AsyncSession):
    dm, eng1, eng2, project = await _setup_data(db)
    user = await create_test_user(db, "ENGINEER", name="Eng1 My")
    user.resource_id = eng1.id
    await db.commit()
    await client.post("/api/v1/auth/login", json={"email": user.email, "password": "TestPass123"})

    resp = await client.get("/api/v1/worklogs/my/export")
    assert resp.status_code == 200
    assert resp.headers["content-type"] == XLSX_CT

    headers, data = _parse_xlsx(resp.content)
    assert headers == ("Date", "Project", "Hours", "Note")  # no Resource column
    assert len(data) == 2


@pytest.mark.asyncio
async def test_export_my_worklogs_no_resource(client: AsyncClient, db: AsyncSession):
    await login_as_role(client, db, "CEO")
    resp = await client.get("/api/v1/worklogs/my/export")
    assert resp.status_code == 403


@pytest.mark.asyncio
async def test_export_my_worklogs_with_filter(client: AsyncClient, db: AsyncSession):
    dm, eng1, eng2, project = await _setup_data(db)
    user = await create_test_user(db, "ENGINEER", name="Eng1 Filter")
    user.resource_id = eng1.id
    await db.commit()
    await client.post("/api/v1/auth/login", json={"email": user.email, "password": "TestPass123"})

    yesterday = (date.today() - timedelta(days=1)).isoformat()
    resp = await client.get(f"/api/v1/worklogs/my/export?start_date={yesterday}&end_date={yesterday}")
    assert resp.status_code == 200
    _, data = _parse_xlsx(resp.content)
    assert len(data) == 1


# ──────────────────────────────────────────────
# GET /api/v1/projects/:id/worklogs/export
# ──────────────────────────────────────────────


@pytest.mark.asyncio
async def test_export_project_worklogs(client: AsyncClient, db: AsyncSession):
    dm, eng1, eng2, project = await _setup_data(db)
    await login_as_role(client, db, "CEO")

    resp = await client.get(f"/api/v1/projects/{project.id}/worklogs/export")
    assert resp.status_code == 200
    assert resp.headers["content-type"] == XLSX_CT

    headers, data = _parse_xlsx(resp.content)
    assert headers == ("Date", "Resource", "Hours", "Note")  # no Project column
    assert len(data) == 3


@pytest.mark.asyncio
async def test_export_project_worklogs_not_found(client: AsyncClient, db: AsyncSession):
    await login_as_role(client, db, "CEO")
    resp = await client.get(f"/api/v1/projects/{uuid.uuid4()}/worklogs/export")
    assert resp.status_code == 404


@pytest.mark.asyncio
async def test_export_project_worklogs_engineer_forbidden(client: AsyncClient, db: AsyncSession):
    """ENGINEER (SELF_ONLY) can't access project-level export."""
    dm, eng1, eng2, project = await _setup_data(db)
    user = await create_test_user(db, "ENGINEER", name="Eng Forbid")
    user.resource_id = eng1.id
    await db.commit()
    await client.post("/api/v1/auth/login", json={"email": user.email, "password": "TestPass123"})

    resp = await client.get(f"/api/v1/projects/{project.id}/worklogs/export")
    assert resp.status_code == 403


@pytest.mark.asyncio
async def test_export_project_worklogs_dm_own_portfolio(client: AsyncClient, db: AsyncSession):
    """DM can export worklogs for projects in own portfolio."""
    dm = await _seed_resource(db, "DM Portfolio")
    eng = await _seed_resource(db, "Eng Port")
    cl = await _seed_client(db)
    own_project = await _seed_project(db, cl.id, dm.id, dm.id, name="OwnProj")
    other_dm = await _seed_resource(db, "Other DM")
    other_project = await _seed_project(db, cl.id, other_dm.id, other_dm.id, name="OtherProj")
    await _seed_assignment(db, own_project.id, eng.id)
    await _seed_assignment(db, other_project.id, eng.id)
    await _seed_worklog(db, eng.id, own_project.id, date.today() - timedelta(days=1))
    await _seed_worklog(db, eng.id, other_project.id, date.today() - timedelta(days=1))
    await db.commit()

    user = await create_test_user(db, "DM", name="DM Port")
    user.resource_id = dm.id
    await db.commit()
    await client.post("/api/v1/auth/login", json={"email": user.email, "password": "TestPass123"})

    # Own project — 200
    resp = await client.get(f"/api/v1/projects/{own_project.id}/worklogs/export")
    assert resp.status_code == 200
    _, data = _parse_xlsx(resp.content)
    assert len(data) == 1

    # Other DM's project — 403
    resp = await client.get(f"/api/v1/projects/{other_project.id}/worklogs/export")
    assert resp.status_code == 403


@pytest.mark.asyncio
async def test_export_project_worklogs_resource_filter(client: AsyncClient, db: AsyncSession):
    dm, eng1, eng2, project = await _setup_data(db)
    await login_as_role(client, db, "CEO")

    resp = await client.get(f"/api/v1/projects/{project.id}/worklogs/export?resource_id={eng1.id}")
    assert resp.status_code == 200
    _, data = _parse_xlsx(resp.content)
    assert len(data) == 2  # only eng1's entries


@pytest.mark.asyncio
async def test_export_filename_includes_date_range(client: AsyncClient, db: AsyncSession):
    await login_as_role(client, db, "CEO")
    start = "2026-01-01"
    end = "2026-01-31"
    resp = await client.get(f"/api/v1/worklogs/export?start_date={start}&end_date={end}")
    assert resp.status_code == 200
    disp = resp.headers["content-disposition"]
    assert start in disp
    assert end in disp
