"""See modules/11-worklog/API.md — Worklog Excel export."""

from datetime import date
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font

from app.modules.worklogs.models import Worklog


def _format_filename(prefix: str, start_date: date | None, end_date: date | None) -> str:
    if start_date and end_date:
        return f"{prefix}_{start_date.isoformat()}_to_{end_date.isoformat()}.xlsx"
    if start_date:
        return f"{prefix}_from_{start_date.isoformat()}.xlsx"
    if end_date:
        return f"{prefix}_until_{end_date.isoformat()}.xlsx"
    return f"{prefix}_all.xlsx"


def generate_worklogs_xlsx(
    rows: list[Worklog],
    *,
    include_resource: bool = True,
    include_project: bool = True,
    start_date: date | None = None,
    end_date: date | None = None,
    filename_prefix: str = "worklogs",
) -> tuple[bytes, str]:
    """Build an xlsx file from worklog rows. Returns (file_bytes, filename)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Worklogs"

    headers: list[str] = ["Date"]
    if include_resource:
        headers.append("Resource")
    if include_project:
        headers.append("Project")
    headers.extend(["Hours", "Note"])

    bold = Font(bold=True)
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = bold

    for row_idx, w in enumerate(rows, 2):
        col = 1
        ws.cell(row=row_idx, column=col, value=w.log_date.isoformat())
        col += 1
        if include_resource:
            ws.cell(row=row_idx, column=col, value=w.resource.name)
            col += 1
        if include_project:
            ws.cell(row=row_idx, column=col, value=w.project.name)
            col += 1
        ws.cell(row=row_idx, column=col, value=float(w.hours))
        col += 1
        ws.cell(row=row_idx, column=col, value=w.note or "")

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = _format_filename(filename_prefix, start_date, end_date)
    return buf.getvalue(), filename
